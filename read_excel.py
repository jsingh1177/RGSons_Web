import openpyxl
import sys

file_path = r"D:\Project\RG Sons\Opening\Opening Stock DHEDEMAU-CS-LKO.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    print(f"Max Column: {sheet.max_column}")
    
    # Print header row specifically
    header = [cell.value for cell in sheet[1]]
    print(f"Header: {header}")

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        row_str = str(row)
        if "14447" in row_str or "113259" in row_str or "4814" in row_str:
            print(f"Row {i} has suspicious value: {row}")
        
except Exception as e:
    print(f"Error: {e}")
